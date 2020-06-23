from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

from webapp.articles.models import ArticleModel


def convert_to_xls(article: ArticleModel) -> bytes:
    """
    Converts Article model to xls book
    """
    book: Workbook = Workbook()
    sheet = book.active

    article_dict = article.__dict__
    tags = ", ".join(article.get_relations("tags"))
    categories = ", ".join(article.get_relations("categories"))
    reference_images = ""
    if article.reference_images:
        reference_images = ", ".join(article.reference_images)
    keys = [
        "unique_id",
        "state",
        "legal_language",
        "citation",
        "cfr40_part280",
        "local_regulation",
        "abstract",
        "updated_date",
        "effective_date",
    ]
    data = [article_dict[key] for key in keys]
    updated_date = article.updated_date
    if updated_date:
        data[-2] = updated_date
    effective_date = article.effective_date
    if effective_date:
        data[-1] = effective_date.strftime("%d/%m/%Y")
    data.insert(-2, tags)
    data.insert(-3, categories)
    data.insert(-4, reference_images)

    header = [
        "Unique ID",
        "State",
        "Regulation",
        "Citation",
        "40CFR 280 Part Federal Rule",
        "Legal Regulation",
        "Abstract",
        "Reference Images",
        "Categories",
        "Tags",
        "Updated Date",
        "Effective Date",
    ]
    sheet.title = article.title
    sheet.append(header)
    sheet.append(data)
    column = 1

    while column < len(data) + 1:
        i = get_column_letter(column)
        sheet.column_dimensions[i].width = 15
        for j in range(1, 13):
            sheet[f"{i}{j}"].alignment = Alignment(
                wrapText=True, vertical="top"
            )
        column += 1
    return save_virtual_workbook(book)
